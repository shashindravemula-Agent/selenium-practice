import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

def update_excel(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1, column = i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row = i, column = j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row = Dict["row"] , column = Dict["col"]).value = new_value
    book.save(file_path)

fruit_name = "Apple"
file_path = "/home/kondas/Downloads/download.xlsx"
newValue = "900"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.find_element(By.ID, "downloadButton").click()

#edit the exel with updated value
update_excel(file_path, fruit_name, "price", newValue)
#upload

file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

toast_locator = (By.XPATH, "//div[@class = 'Toastify__toast-body']/div [2]")
WebDriverWait(driver, 5).until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
Actual_price = driver.find_element(By.XPATH, "//div[text() = '"+fruit_name+"']/parent::div/parent::div/div[@id = 'cell-"+price_column+"-undefined']").text
assert Actual_price == newValue