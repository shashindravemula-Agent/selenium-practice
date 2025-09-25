from openpyxl import Workbook

# Create a workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Fruits"

# Add headers
sheet["A1"] = "Fruit"
sheet["B1"] = "Price"

# Add some data
sheet.append(["Apple", 100])
sheet.append(["Banana", 50])
sheet.append(["Mango", 80])

wb.save("fruits.xlsx")
print("Sample Excel created ✅")



from openpyxl import load_workbook

# Load the workbook
wb = load_workbook("fruits.xlsx")
sheet = wb["Fruits"]

# Update price of Banana
for row in sheet.iter_rows(min_row=2, values_only=False):
    if row[0].value == "Banana":
        row[1].value = 60  # new price

wb.save("fruits_updated.xlsx")
print("Excel updated ✅")




from selenium import webdriver
from selenium.webdriver.common.by import By
import time

driver = webdriver.Chrome()
driver.get("https://the-internet.herokuapp.com/upload")

# Locate file input & upload
file_input = driver.find_element(By.ID, "file-upload")
file_input.send_keys("/home/kondas/PycharmProjects/selenium-practice/fruits_updated.xlsx")

# Click upload button
driver.find_element(By.ID, "file-submit").click()

time.sleep(3)

# Verify upload success
success = driver.find_element(By.TAG_NAME, "h3").text
print("Upload Status:", success)

driver.quit()
